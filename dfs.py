#%% Imports
import os
import numpy as np
import pandas as pd
import networkx as nx
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill
from contextlib import contextmanager

#%% Classes, Function, Constants
class InvalidSortingMethod(Exception):
    def __init__(self,message):
        super().__init__(message)
        
class InvalidPathSearchMethod(Exception):
    def __init__(self,message):
        super().__init__(message)
class NodeNotFound(Exception):
    def __init__(self,message):
        super().__init__(message)

class Node():
    def __init__(self,name,location):
        self.name = name
        self.location = location
        self.neighbors = []

    def __repr__(self):
        return f"""Name: {self.name}, Location: {self.location}"""
    
    def get_manhattan(self,node):
        """Solves for the Manhattan distance
        between this node and a specified node name."""
        if node.name == self.name:
            return 0
        xn,yn = node.location
        xo,yo = self.location
        return abs(xn-xo) + abs(yn-yo)
    
    def filter_neighbor(self,node):
        """filter function used to find th
        neighbor/s of this node."""
        manhattan = self.get_manhattan(node)
        if manhattan == 0:
            return False
        if manhattan == 1:
            return True
        return False
    
    def get_neighbors(self,nodes):
        """Returns the neighbors of this node."""
        self.neighbors = list(filter(self.filter_neighbor,nodes))
        return self.neighbors

class Graph():
    def __init__(self,nodes):
        self.nodes = nodes
        self.create_adjacency_matrix(self.nodes)
        self.graph = nx.from_numpy_array(self.adjacency_matrix, create_using=nx.MultiGraph())

    def create_adjacency_matrix(self,nodes):
        """Create the adjacency matrix needed to create the graph.
        This is used in conjuction with the self.draw function."""
        self.node_count = len(nodes)
        self.adjacency_matrix = np.zeros((self.node_count,self.node_count))
        for node in nodes:
            neighbors = node.get_neighbors(nodes)
            for neighbor in neighbors:
                self.adjacency_matrix[neighbor.name,node.name] = 1
        return self.adjacency_matrix
    
    def draw(self,**kwargs):
        """Draw the network structure."""
        fig = plt.figure()
        pos = nx.nx_pydot.pydot_layout(self.graph,prog="twopi")
        nx.draw_networkx(self.graph,pos = pos,font_size = 5,**kwargs)
        plt.show()
        return fig
    
    def get_node(self,name):
        """Get the node object from a given node name."""
        filtered = list(filter(lambda node:node.name == name,self.nodes))
        if len(filtered) == 0:
            raise NodeNotFound(f"Node: {name} not found in the node list.")
        return filtered[0]
    
    def manhattan_sort(self,end,neighbors,method = 'dfs'):
        """
        A naive sorting for neighbors before adding to
        the stack/queue. Node with east Manhattan distance is placed
        last in dfs stack. Node with least Manhattan distance is place first
        in = bfs queue. this can be improved by getting the actual path
        distance. This would require a recursive approach though.

        Parameters
        ----------
        end : int
            End node name
        neighbors : list([Node])
            A list containg Node objects that are the immediate neighbor of the
            current node.
        method : str, optional
            Specifies the path search algorithm used.\n
            'dfs' - Depth-First Search prioritizes search for the deepest
                    path that can be found for each neighboring node.\n
            'bfs' - Breadth-First Search prioritizes searching through all
                    of the closest nodes before moving to the next order
                    neighbors\n
            The default is "dfs".
        Returns
        -------
        sorted_neighbors : list([Node])
            Contains list of Node objects listed according to the selected path
            search method.
        """
        
        end_node = self.get_node(end)
        if method == "dfs":
            sorted_neighbors = sorted(neighbors,
                                      key = lambda node: node.get_manhattan(end_node),
                                      reverse = True)
        elif method == "bfs":
            sorted_neighbors = sorted(neighbors,
                                      key = lambda node: node.get_manhattan(end_node),
                                      reverse = False)
        return sorted_neighbors

    def path_search(self,start,end,method = "dfs",neighbor_sort_method = "manhattan"):
        """
        Performs Depth-First Search (DFS) or a Breadth-First Search (BFS)
        to find a path between the start and end node.
        

        Parameters
        ----------
        start : int
            Starting node name
        end : int
            Ending node name
        method : str, optional
            Specifies the path search algorithm used.\n
            'dfs' - Depth-First Search prioritizes search for the deepest
                    path that can be found for each neighboring node.\n
            'bfs' - Breadth-First Search prioritizes searching through all
                    of the closest nodes before moving to the next order
                    neighbors\n
            The default is "dfs".
        neighbor_sort_method : str or None, optional
            Specifies the immediate neighbor sorting method before being
            appended into a stack/queue.\n
            'manhattan' - sort according to the least Manhattan distance\n
            'random' - sort immediate neighbor randomly\n
            None - append neighbors as is as how they were listed\n
            The default is "manhattan".

        Raises
        ------
        InvalidPathSearchMethod
            Path search method chosen is invalid. Select from 'dfs' or 'bfs'.
        InvalidSortingMethod
            Neighbor sorting method is invalid. Select from 'manhattan',
            'random' or None.

        Returns
        -------
        parent_map : dict
            Returns a dictionary mapping the child node to the parent
            used to discover itself. Used to trace back the path taken
            by the search algorithm.
        """
        
        explored = [start]
        stack = [start] #appropriate name is 'queue' if bfs is used
        parent_map = {start:None}
        while len(stack)>0:
            if method == "dfs":
                node = stack.pop()
            elif method == "bfs":
                node = stack.pop(0)
            else:
                raise InvalidPathSearchMethod(f"Invalid path search method ({method}). Choose dfs or bfs")
            
            if node == end:
                return parent_map
            neighbors = np.array(self.get_node(node).neighbors)
            
            if neighbor_sort_method == "manhattan":
                #sort according to least manhattan distance from end node
                neighbors = self.manhattan_sort(end,neighbors,method)
            elif neighbor_sort_method == "random":
                #randomly shuffle the values
                np.random.shuffle(neighbors)
            elif neighbor_sort_method is None:
                #just pass the neighbors as is
                pass
            else:
                raise InvalidSortingMethod(
                    f"""Invalid neighbor sorting method ({neighbor_sort_method}). Choose manhattan, random, or None.""")

            for neighbor in neighbors:
                if not neighbor.name in explored:
                    stack.append(neighbor.name)
                    explored.append(neighbor.name)
                    parent_map[neighbor.name]=node
                    if node == end:
                        return parent_map

    def get_path(self,start,end,method = "dfs",neighbor_sort_method = "manhattan"):
        """Unfolds the parent map to return the path taken.
        The path returned is not quaranteed as the shortest path."""
        parent_map = self.path_search(start,end,method,neighbor_sort_method = neighbor_sort_method)
        current_node = end
        path = [current_node]
        while True:
            current_node = parent_map[current_node]
            if not current_node:
                path.append(start)
                return path
            path.append(current_node)
      
def close_excel():
    try:
        os.system("TASKKILL /F /IM EXCEL.EXE")
    except Exception as e:
        print(e)
def open_excel(file_name):
    try:
        os.system(f"start EXCEL.EXE {file_name}")
    except Exception as e:
        print(e)

def string_length(value):
    if not value:
        return 0
    if type(value) == str:
        return len(value)
    else:
        return len(str(value))
    
def custom_value(value):
    if value == "#":
        return -1
    elif value in path:
        return 1
    else:
        return 0

def heatmap_path(start,end,data_frame,ax,graph):
    """
    Show the path taken as heatmap. Node names won't be
    visible due to size restriction.

    Parameters
    ----------
    start : int
        Starting node name
    end : int
        Ending node name
    data_frame : pandas.DataFrame
        A data frame containing the node names and the walls represented
        as "#". Used to draw the maze.
    ax : matplotlib.axes._axes.Axes
        An axis to draw the heatmap.
    graph : Graph
        A Graph object of the maze. Use to find the location of the start
        and end nodes.
    Returns
    -------
    None.

    """
    annotation = np.full(data_frame.shape,"")
    xs,ys = graph.get_node(start).location
    xe,ye = graph.get_node(end).location
    annotation[xs,ys] = "S"
    annotation[xe,ye] = "E"
    annot_kws = {"color":"white","verticalalignment":"center",
                 "horizontalalignment":"center",
                 "fontweight":"bold"}
    
    maze_heatmap = data_frame.applymap(custom_value)
    colors = ListedColormap(["black","red","#00FF0080"])
    sns.heatmap(maze_heatmap,cbar=False,cmap =colors,
                annot=annotation,annot_kws = annot_kws,
                fmt="s",ax = ax)


RED_FILL = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
GREEN_FILL = PatternFill(start_color='00b050',
                   end_color='00b050',
                   fill_type='solid')

@contextmanager
def WriterContext(writer_obj):
    try:
        yield writer_obj
    finally:
        if len(writer_obj.sheets) > 0:
            writer_obj.close()

def draw_path_excel(path,data_frame,sheet_name,writer):
    """
    Generates an Excel file highlighting the path in the maze.
    
    Parameters
    ----------
    path : list([int])
        A list of node names tracing back the path from the end node to the
        start node.
    data_frame : pandas.DataFrame
        A data frame containing the node names and the walls represented
        as "#". Used to draw the maze.
    sheet_name : str
        Sheet name to save the maze solution to.
    writer : pandas.io.excel._openpyxl.OpenpyxlWriter
        Excel writer object.

    Returns
    -------
    None.

    """
        
    data_frame.to_excel(writer_obj,sheet_name = sheet_name, index=True)
    sheet = writer.book[sheet_name]
    for idx, column in enumerate(sheet.columns):
        col = get_column_letter(idx + 1) #index starts at 1
        values = []
        for cell_idx,cell in enumerate(sheet[col]):
            values.append(cell.value)
            cell.alignment = Alignment(horizontal = "center", vertical = "center")
            if idx > 0 and cell_idx > 0: #do not include index
                if cell.value != "#" and not cell.value in path:
                    cell.fill = RED_FILL
                if cell.value in path:
                    cell.fill = GREEN_FILL
                    
        max_length = max(map(string_length,values))
        sheet.column_dimensions[col].width = (max_length + 2)*1
        
        for i in range(len(list(sheet.rows))):
            sheet.row_dimensions[i+1].height = (max_length + 2)*5
#%% Search Path

#Load maze from excel file.
sheet_name = "maze_101by101" #maze_20by20 or maze_101by101
maze = pd.read_excel("maze.xlsx",sheet_name=sheet_name)
maze = maze.drop("Row",axis = 1)
maze_array  =np.array(maze)
maze_labeled  = maze_array.copy()

#Construct the node objects
nodes = []
node_count = 0
for i in range(maze.shape[0]):
    for j  in range(maze.shape[1]):
        element = maze_array[i,j]
        if element == "ST":
            start = (i,j)
        elif element == "FN":
            end = (i,j)
        if element != "#":
            location = (i,j)
            nodes.append(Node(node_count,location))
            maze_labeled[i,j] = node_count
            node_count += 1
print(f"Total number of nodes: {node_count}")

file_name = "path.xlsx"
maze_path = pd.DataFrame(maze_labeled,index = maze.index,columns = maze.columns)
writer_obj = pd.ExcelWriter(file_name, engine='openpyxl')

mygraph = Graph(nodes)
methods = ["bfs","dfs"]
starts = np.random.choice(range(0,3500),5)
ends = np.random.choice(range(4000,5007),5)

rows = len(methods)
cols = len(starts)
fig,axes = plt.subplots(nrows = rows, ncols = cols,figsize = (15,7.5))
axes = np.ravel(axes)
plot_counter = 0

plt.ioff()
plt.close("all")
close_excel()
with WriterContext(writer_obj) as writer:
    for method in methods:
        print(f"Current Method: {method}")
        for i,(start,end) in enumerate(zip(starts,ends)):
            path = mygraph.get_path(start,end,method = method,neighbor_sort_method = "manhattan")
            draw_path_excel(path,maze_path,f"path_{method}_{start}_{end}",writer)
            print(f"{i+1}. Path found for {start} to {end}")
            
            ax = axes[plot_counter]
            heatmap_path(start,end,maze_path,ax,mygraph)
            plot_counter += 1
            ax.set_title(f"{method}_{start}_{end}")
            ax.set_aspect(1)
            
open_excel(file_name)
plt.tight_layout()
fig.savefig("paths.png",dpi = 600)
fig.show()






        


